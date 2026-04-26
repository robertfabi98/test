#!/usr/bin/env python3
"""
Genera un file Excel di contabilità giornaliera dal messaggio standard.
Uso: python scripts/contabilita.py messaggio.txt
     oppure: echo "..." | python scripts/contabilita.py
"""
import re
import os
import sys

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("Errore: installa openpyxl con: pip install openpyxl")
    sys.exit(1)

MONTHS_IT = {
    "GENNAIO": 1, "FEBBRAIO": 2, "MARZO": 3, "APRILE": 4,
    "MAGGIO": 5, "GIUGNO": 6, "LUGLIO": 7, "AGOSTO": 8,
    "SETTEMBRE": 9, "OTTOBRE": 10, "NOVEMBRE": 11, "DICEMBRE": 12,
}
MONTHS_NAME = {v: k.capitalize() for k, v in MONTHS_IT.items()}


def parse_message(text: str) -> dict:
    data = {"day": None, "month": None, "year": None,
            "entrate": [], "spese": [], "stipendi": [], "borselli": []}
    current = None
    for raw in text.splitlines():
        line = raw.strip()
        if not line:
            continue
        m = re.match(r'Data[:\s]+(\d{1,2})\s+([A-ZÀÈÌÒÙ]+)\s+(\d{4})', line, re.IGNORECASE)
        if m:
            data["day"] = int(m.group(1))
            data["month"] = MONTHS_IT.get(m.group(2).upper(), 1)
            data["year"] = int(m.group(3))
            continue
        if re.search(r'ENTRATE', line, re.IGNORECASE):
            current = "entrate"; continue
        if re.search(r'SPESE', line, re.IGNORECASE):
            current = "spese"; continue
        if re.search(r'STIPENDI', line, re.IGNORECASE):
            current = "stipendi"; continue
        if re.search(r'BORSELLI', line, re.IGNORECASE):
            current = "borselli"; continue
        m2 = re.match(r'(.+?)\s*(?:—>|-->|->|>)\s*([\d.,]+)', line)
        if m2 and current:
            label = m2.group(1).strip().lstrip('-').strip()
            val = float(m2.group(2).replace(',', '.'))
            data[current].append({"label": label, "value": val})
    return data


def thin_border():
    s = Side(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)


def make_excel(data: dict, out_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 1   # spacer column

    C_TITLE    = "1A252F"
    C_ENTRATE  = "1E8449"
    C_SPESE    = "C0392B"
    C_STIP     = "D35400"
    C_BORSELLI = "1A5276"
    C_RIEPILOGO = "6C3483"
    C_ALT      = "F2F3F4"
    C_TOTAL    = "D6EAF8"

    row = 1

    def section_title(text, color):
        nonlocal row
        ws.merge_cells(f"A{row}:B{row}")
        c = ws.cell(row=row, column=1, value=text)
        c.font = Font(bold=True, color="FFFFFF", size=12)
        c.fill = PatternFill("solid", fgColor=color)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 24
        row += 1

    def col_header():
        nonlocal row
        for col, val in [(1, "Voce"), (2, "Importo")]:
            c = ws.cell(row=row, column=col, value=val)
            c.font = Font(bold=True, size=9, color="555555")
            c.fill = PatternFill("solid", fgColor="D5D8DC")
            c.border = thin_border()
            c.alignment = Alignment(horizontal="right" if col == 2 else "left", vertical="center")
        ws.row_dimensions[row].height = 16
        row += 1

    def entry_row(label, value, alt=False):
        nonlocal row
        bg = C_ALT if alt else "FFFFFF"
        c1 = ws.cell(row=row, column=1, value=label)
        c2 = ws.cell(row=row, column=2, value=value)
        c2.number_format = '#,##0.00 "€"'
        for c in (c1, c2):
            c.fill = PatternFill("solid", fgColor=bg)
            c.border = thin_border()
            c.alignment = Alignment(vertical="center")
        c2.alignment = Alignment(horizontal="right", vertical="center")
        ws.row_dimensions[row].height = 16
        row += 1

    def total_row(label, value):
        nonlocal row
        c1 = ws.cell(row=row, column=1, value=label)
        c2 = ws.cell(row=row, column=2, value=value)
        c2.number_format = '#,##0.00 "€"'
        for c in (c1, c2):
            c.font = Font(bold=True)
            c.fill = PatternFill("solid", fgColor=C_TOTAL)
            c.border = thin_border()
            c.alignment = Alignment(vertical="center")
        c2.alignment = Alignment(horizontal="right", vertical="center")
        ws.row_dimensions[row].height = 18
        row += 1
        return value

    def spacer():
        nonlocal row
        ws.row_dimensions[row].height = 8
        row += 1

    # ── TITLE ──────────────────────────────────────────────────────────────
    mn = MONTHS_NAME.get(data["month"], "")
    ws.merge_cells(f"A{row}:B{row}")
    c = ws.cell(row=row, column=1,
                value=f"REPORT CONTABILE  ·  {data['day']:02d} {mn.upper()} {data['year']}")
    c.font = Font(bold=True, color="FFFFFF", size=14)
    c.fill = PatternFill("solid", fgColor=C_TITLE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 30
    row += 1
    spacer()

    # ── ENTRATE ────────────────────────────────────────────────────────────
    section_title("💰  ENTRATE", C_ENTRATE)
    col_header()
    tot_e = 0
    for i, e in enumerate(data["entrate"]):
        entry_row(e["label"], e["value"], alt=i % 2 == 1)
        tot_e += e["value"]
    if not data["entrate"]:
        entry_row("(nessuna entrata)", 0)
    total_row("TOTALE ENTRATE", tot_e)
    spacer()

    # ── SPESE ──────────────────────────────────────────────────────────────
    section_title("💸  SPESE", C_SPESE)
    col_header()
    tot_s = 0
    for i, e in enumerate(data["spese"]):
        entry_row(e["label"], e["value"], alt=i % 2 == 1)
        tot_s += e["value"]
    if not data["spese"]:
        entry_row("(nessuna spesa)", 0)
    total_row("TOTALE SPESE", tot_s)
    spacer()

    # ── STIPENDI ───────────────────────────────────────────────────────────
    section_title("👥  STIPENDI", C_STIP)
    col_header()
    tot_stip = 0
    for i, e in enumerate(data["stipendi"]):
        entry_row(e["label"], e["value"], alt=i % 2 == 1)
        tot_stip += e["value"]
    if not data["stipendi"]:
        entry_row("(nessun stipendio)", 0)
    total_row("TOTALE STIPENDI", tot_stip)
    spacer()

    # ── BORSELLI ───────────────────────────────────────────────────────────
    section_title("👜  BORSELLI", C_BORSELLI)
    col_header()
    tot_b = 0
    for i, e in enumerate(data["borselli"]):
        entry_row(e["label"], e["value"], alt=i % 2 == 1)
        tot_b += e["value"]
    if not data["borselli"]:
        entry_row("(nessun borsello)", 0)
    total_row("TOTALE BORSELLI", tot_b)
    spacer()

    # ── RIEPILOGO ──────────────────────────────────────────────────────────
    section_title("📊  RIEPILOGO", C_RIEPILOGO)
    entry_row("Totale Entrate",  tot_e)
    entry_row("Totale Spese",    tot_s,    alt=True)
    entry_row("Totale Stipendi", tot_stip)
    entry_row("Totale Borselli", tot_b,    alt=True)
    spacer()

    netto = tot_e - tot_s - tot_stip
    nc = ws.cell(row=row, column=1, value="RISULTATO (Entrate - Spese - Stipendi)")
    nv = ws.cell(row=row, column=2, value=netto)
    nv.number_format = '#,##0.00 "€"'
    col = "1E8449" if netto >= 0 else "C0392B"
    for c in (nc, nv):
        c.font = Font(bold=True, color="FFFFFF", size=12)
        c.fill = PatternFill("solid", fgColor=col)
        c.border = thin_border()
        c.alignment = Alignment(vertical="center")
    nv.alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[row].height = 26

    wb.save(out_path)
    print(f"✅  Salvato: {out_path}")


def main():
    if len(sys.argv) > 1 and os.path.exists(sys.argv[1]):
        with open(sys.argv[1], encoding="utf-8") as f:
            text = f.read()
    elif len(sys.argv) > 1:
        text = " ".join(sys.argv[1:])
    else:
        text = sys.stdin.read()

    data = parse_message(text)
    if not data["day"]:
        print("Errore: data non trovata. Includi una riga come: Data: 25 APRILE 2026")
        sys.exit(1)

    mn = MONTHS_NAME.get(data["month"], "Mese")
    out_dir = os.path.join("contabilita", str(data["year"]), mn)
    os.makedirs(out_dir, exist_ok=True)
    out_file = os.path.join(out_dir, f"{data['day']:02d}_{mn}_{data['year']}.xlsx")
    make_excel(data, out_file)


if __name__ == "__main__":
    main()
