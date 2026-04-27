import os
from typing import Optional

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

from parser import DayRecord, MESI_IT_INV

_C = {
    'dark':     '1F4E79',
    'white':    'FFFFFF',
    'entrate':  'E2EFDA',
    'spese':    'FCE4D6',
    'stipendi': 'FFF2CC',
    'borselli': 'DAEEF3',
    'summary':  'D6DCE4',
    'alt':      'F5F5F5',
}


def _border():
    s = Side(border_style='thin', color='BFBFBF')
    return Border(left=s, right=s, top=s, bottom=s)


def _fmt(cell, value=None, *, bold=False, size=10, fg='000000',
         bg: Optional[str] = None, align='left', num_fmt: Optional[str] = None):
    if value is not None:
        cell.value = value
    cell.font = Font(bold=bold, size=size, color=fg)
    cell.fill = PatternFill(start_color=bg or 'FFFFFF', end_color=bg or 'FFFFFF', fill_type='solid')
    cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
    cell.border = _border()
    if num_fmt:
        cell.number_format = num_fmt


def _section(ws: Worksheet, row: int, title: str, data: dict, bg: str) -> int:
    ws.merge_cells(f'A{row}:B{row}')
    _fmt(ws[f'A{row}'], title, bold=True, size=10, fg=_C['white'], bg=_C['dark'], align='center')
    ws.row_dimensions[row].height = 18
    row += 1

    if data:
        for i, (k, v) in enumerate(data.items()):
            row_bg = bg if i % 2 == 0 else _C['alt']
            _fmt(ws[f'A{row}'], k, size=10, bg=row_bg)
            _fmt(ws[f'B{row}'], v, size=10, bg=row_bg, align='right', num_fmt='#,##0.00 €')
            row += 1
    else:
        ws.merge_cells(f'A{row}:B{row}')
        _fmt(ws[f'A{row}'], '(nessuna voce)', size=9, fg='808080', bg='F9F9F9', align='center')
        row += 1

    total = sum(data.values()) if data else 0.0
    _fmt(ws[f'A{row}'], f'TOTALE {title}', bold=True, size=10, bg=_C['summary'])
    _fmt(ws[f'B{row}'], total, bold=True, size=10, bg=_C['summary'], align='right', num_fmt='#,##0.00 €')
    return row + 2


def _day_sheet(wb: openpyxl.Workbook, r: DayRecord) -> None:
    name = f'{r.date.day:02d}'
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(title=name)
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 16

    month = MESI_IT_INV.get(r.date.month, '')
    ws.merge_cells('A1:B1')
    _fmt(ws['A1'], f'CONTABILITÀ — {r.date.day} {month} {r.date.year}',
         bold=True, size=12, fg=_C['white'], bg=_C['dark'], align='center')
    ws.row_dimensions[1].height = 24

    row = 3
    row = _section(ws, row, 'ENTRATE',  r.entrate,  _C['entrate'])
    row = _section(ws, row, 'SPESE',    r.spese,    _C['spese'])
    row = _section(ws, row, 'STIPENDI', r.stipendi, _C['stipendi'])
    row = _section(ws, row, 'BORSELLI', r.borselli, _C['borselli'])

    ws.merge_cells(f'A{row}:B{row}')
    _fmt(ws[f'A{row}'], 'RIEPILOGO GIORNALIERO', bold=True, size=11,
         fg=_C['white'], bg=_C['dark'], align='center')
    ws.row_dimensions[row].height = 20
    row += 1

    for label, val, bg in [
        ('Totale Entrate',  r.totale_entrate,  _C['entrate']),
        ('Totale Spese',    r.totale_spese,    _C['spese']),
        ('Totale Stipendi', r.totale_stipendi, _C['stipendi']),
        ('Totale Borselli', r.totale_borselli, _C['borselli']),
    ]:
        _fmt(ws[f'A{row}'], label, bold=True, size=10, bg=bg)
        _fmt(ws[f'B{row}'], val,   bold=True, size=10, bg=bg, align='right', num_fmt='#,##0.00 €')
        row += 1

    saldo = r.saldo
    _fmt(ws[f'A{row}'], 'SALDO NETTO', bold=True, size=11, fg=_C['white'], bg=_C['dark'])
    _fmt(ws[f'B{row}'], saldo, bold=True, size=11, fg=_C['white'], bg=_C['dark'],
         align='right', num_fmt='#,##0.00 €')


def _summary_sheet(wb: openpyxl.Workbook) -> None:
    TAB = 'RIEPILOGO'
    if TAB in wb.sheetnames:
        del wb[TAB]
    ws = wb.create_sheet(title=TAB, index=0)

    headers  = ['Giorno', 'Entrate', 'Spese', 'Stipendi', 'Borselli', 'Saldo']
    hdr_bgs  = [_C['summary'], _C['entrate'], _C['spese'], _C['stipendi'], _C['borselli'], _C['summary']]
    col_widths = [10, 14, 14, 14, 14, 14]

    for col_letter, w in zip('ABCDEF', col_widths):
        ws.column_dimensions[col_letter].width = w

    ws.merge_cells('A1:F1')
    _fmt(ws['A1'], 'RIEPILOGO MENSILE', bold=True, size=13,
         fg=_C['white'], bg=_C['dark'], align='center')
    ws.row_dimensions[1].height = 24

    for col, (h, bg) in enumerate(zip(headers, hdr_bgs), 1):
        _fmt(ws.cell(2, col), h, bold=True, size=10, bg=bg, align='center')

    totals = [0.0] * 5  # entrate, spese, stipendi, borselli, saldo
    data_row = 3

    for sname in wb.sheetnames:
        if sname == TAB:
            continue
        ws_d = wb[sname]
        vals: dict = {}
        for r in ws_d.iter_rows(values_only=True):
            if not r[0]:
                continue
            lbl = str(r[0]).upper()
            v = r[1]
            if   'TOTALE ENTRATE'  in lbl and v is not None: vals['e'] = v
            elif 'TOTALE SPESE'    in lbl and v is not None: vals['s'] = v
            elif 'TOTALE STIPENDI' in lbl and v is not None: vals['t'] = v
            elif 'TOTALE BORSELLI' in lbl and v is not None: vals['b'] = v
            elif 'SALDO NETTO'     in lbl and v is not None: vals['n'] = v

        row_vals = [
            sname,
            vals.get('e', 0.0),
            vals.get('s', 0.0),
            vals.get('t', 0.0),
            vals.get('b', 0.0),
            vals.get('n', 0.0),
        ]
        row_bg = _C['alt'] if data_row % 2 == 0 else 'FFFFFF'

        for col, v in enumerate(row_vals, 1):
            cell = ws.cell(data_row, col)
            if col == 1:
                _fmt(cell, v, size=10, bg=row_bg, align='center')
            else:
                _fmt(cell, v, size=10, bg=row_bg, align='right', num_fmt='#,##0.00 €')
                totals[col - 2] += float(v) if v else 0.0

        data_row += 1

    _fmt(ws.cell(data_row, 1), 'TOTALE', bold=True, size=10, bg=_C['summary'], align='center')
    for i, total in enumerate(totals, 2):
        _fmt(ws.cell(data_row, i), total, bold=True, size=10,
             bg=_C['summary'], align='right', num_fmt='#,##0.00 €')


def generate_excel(record: DayRecord, output_dir: str = 'contabilita') -> str:
    month = MESI_IT_INV.get(record.date.month, f'{record.date.month:02d}')
    year_dir = os.path.join(output_dir, str(record.date.year))
    os.makedirs(year_dir, exist_ok=True)
    path = os.path.join(year_dir, f'contabilita_{month}_{record.date.year}.xlsx')

    wb = openpyxl.load_workbook(path) if os.path.exists(path) else openpyxl.Workbook()
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    _day_sheet(wb, record)
    _summary_sheet(wb)
    wb.save(path)
    return path
